# ******************************************************************************
#  This file is part of pymdu.                                                 *
#                                                                              *
#  pymdu is free software: you can redistribute it and/or modify               *
#  it under the terms of the GNU General Public License as published by        *
#  the Free Software Foundation, either version 3 of the License, or           *
#  (at your option) any later version.                                         *
#                                                                              *
#  pymdu is distributed in the hope that it will be useful,                    *
#  but WITHOUT ANY WARRANTY; without even the implied warranty of              *
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the               *
#  GNU General Public License for more details.                                *
#                                                                              *
#  You should have received a copy of the GNU General Public License           *
#  along with pymdu.  If not, see <https://www.gnu.org/licenses/>.             *
# ******************************************************************************
import json
import os
import platform
from datetime import datetime

import fitz
import folium
import geopandas as gpd
import geopy
import matplotlib.cm as cm
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import pdfkit
import requests
import seaborn as sns
import shapely
from geopy.distance import geodesic
from joblib import delayed, Parallel
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from pyephem_sunpath.sunpath import sunpos
from scipy.spatial import distance
from shapely import Point
from windrose import WindroseAxes

from pymdu._typing import FilePath
from pymdu.collect.GlobalVariables import TEMP_PATH


class Diagnostic:
    def __init__(
        self,
        output_path: FilePath = None,
        address: str = '8 rue isabelle autissier, 17000 Lagord',
        year: int = 2018,
    ):
        super().__init__()

        self.address: str = address
        self.year: int = year
        self.output_path: FilePath = output_path if output_path else TEMP_PATH
        self.gdf: gpd.GeoDataFrame = None
        self.zonage_plu = None
        self.lat: float = None
        self.lon: float = None
        self.parcelle_commnune = None
        self.parcelle_number = None
        self.parcelle_bbox = None
        self.altitude: float = None
        self.department: str = None
        self.insee_code = None
        self.parcelle_surface = None
        self.point = None
        self.isochrone = None
        self.pluviometrie_moyenne_mensuelle = None
        self.temperature_moyenne_mensuelle = None
        self.radon = None
        self.StationLaPlusProche = None
        self.StationsDuDepartement = None
        self.cours_d_eau = None
        self.mvtTerrain = None
        self.weather_data = None
        self.path_windrose: FilePath = None
        self.path_zonage_inondations = None
        self.path_radon: FilePath = None
        self.path_mvt_terrain: FilePath = None
        self.path_retrait_argile: FilePath = None
        self.rues_delimitantes = None
        self.path_course_soleil: FilePath = None
        self.path_zonage_sismique: FilePath = None
        self.path_remontee_nappe: FilePath = None
        self.path_liste_cours_deau: FilePath = None
        self.path_isochrones: FilePath = None
        self.path_bassin_versant: FilePath = None
        self.path_map_commerces: FilePath = None
        self.path_map_equipements_culturels_loisirs: FilePath = None
        self.path_map_etablissements_scolaires: FilePath = None
        self.path_map_service_sante: FilePath = None
        self.path_pluviometrie_moyenne_mensuelle: FilePath = None
        self.path_temperature_moyenne_mensuelle: FilePath = None
        self.path_localisation_du_site: FilePath = None
        self.path_parcelle_cadastrale: FilePath = None
        self.path_zonage_plu: FilePath = None
        self.path_rues_delimitantes: FilePath = None

        self.path_altitude: FilePath = None

        self.info_windrose = None
        self.info_zonage_inondations = None
        self.info_radon = None
        self.info_mvt_terrain = None
        self.info_retrait_argile = None
        self.rues_delimitantes = None
        self.info_course_soleil = None
        self.info_zonage_sismique = None
        self.info_remontee_nappe = None
        self.info_liste_cours_deau = None
        self.info_isochrones = None
        self.info_bassin_versant = None
        self.info_map_commerces = None
        self.info_map_equipements_culturels_loisirs = None
        self.info_map_etablissements_scolaires = None
        self.info_map_service_sante = None
        self.info_pluviometrie_moyenne_mensuelle = None
        self.info_temperature_moyenne_mensuelle = None

    def run(self, filename: str = 'Diagnostic.xlsx'):
        configuration_diagnostic = [
            {
                'THEMATIQUE': 'Localisation du site',
                'INFORMATION': 'Coordonnées GPS',
                'SOURCE': 'https://adresse.data.gouv.fr/api-doc/adresse',
                'DESCRIPTION': 'Coordonnées GPS',
                'FORMAT': 'STR',
                'UNIT': '',
                'FUNCTION': 'get_localisation_du_site',
                'PATH': 'path_localisation_du_site',
                'VARIABLES': ['lon', 'lat'],
                'LIEN UTILE': 'https://adresse.data.gouv.fr/',
            },
            {
                'THEMATIQUE': 'Localisation du site',
                'INFORMATION': 'Parcelle cadastrale',
                'SOURCE': 'https://api.gouv.fr/les-api/api_carto_cadastre',
                'DESCRIPTION': 'N° de parcelle',
                'FORMAT': 'FLOAT',
                'UNIT': '',
                'FUNCTION': 'get_parcelle_cadastrale',
                'PATH': 'path_parcelle_cadastrale',
                'VARIABLES': ['parcelle_number'],
                'LIEN UTILE': 'https://cadastre.data.gouv.fr/',
            },
            {
                'THEMATIQUE': 'Localisation du site',
                'INFORMATION': 'Superficie de la parcelle',
                'SOURCE': 'https://api.gouv.fr/les-api/api_carto_cadastre',
                'DESCRIPTION': 'm2 de parcelle',
                'FORMAT': 'FLOAT',
                'UNIT': 'm2',
                'FUNCTION': 'get_parcelle_cadastrale',
                'PATH': 'path_parcelle_cadastrale',
                'VARIABLES': ['parcelle_surface'],
                'LIEN UTILE': 'https://cadastre.data.gouv.fr/',
            },
            {
                'THEMATIQUE': 'Localisation du site',
                'INFORMATION': 'Localisation du site sur zonage PLU',
                'SOURCE': 'https://www.geoportail-urbanisme.gouv.fr/api/',
                'DESCRIPTION': 'Type de zone PLU',
                'FORMAT': 'STR',
                'UNIT': '',
                'FUNCTION': 'get_zonage_plu',
                'PATH': 'path_zonage_plu',
                'VARIABLES': ['zonage_plu'],
                'LIEN UTILE': 'https://docurba.beta.gouv.fr/',
            },
            {
                'THEMATIQUE': 'Climatologie',
                'INFORMATION': 'Moyennes mensuelles de la température de la station météo la plus proche',
                'SOURCE': 'https://meteo.data.gouv.fr/',
                'DESCRIPTION': 'Schéma mensuelle de température',
                'FORMAT': 'IMG',
                'UNIT': '°C',
                'FUNCTION': 'get_temperature_moyenne_mensuelle',
                'PATH': 'path_temperature_moyenne_mensuelle',
                'VARIABLES': ['path_temperature_moyenne_mensuelle'],
                'LIEN UTILE': 'https://meteo.data.gouv.fr/',
            },
            {
                'THEMATIQUE': 'Climatologie',
                'INFORMATION': 'Moyennes mensuelles de la pluviométrie sur la station météo la plus proche',
                'SOURCE': 'https://meteo.data.gouv.fr/',
                'DESCRIPTION': 'Schéma mensuelle de la pluviométrie',
                'FORMAT': 'IMG',
                'UNIT': 'mm',
                'FUNCTION': 'get_pluviometrie_moyenne_mensuelle',
                'PATH': 'path_pluviometrie_moyenne_mensuelle',
                'VARIABLES': ['path_pluviometrie_moyenne_mensuelle'],
                'LIEN UTILE': 'https://meteo.data.gouv.fr/',
            },
            {
                'THEMATIQUE': 'Climatologie',
                'INFORMATION': 'Vents dominants: orientation et puissance',
                'SOURCE': 'https://meteo.data.gouv.fr/',
                'DESCRIPTION': 'Rose des vents saisonnière',
                'FORMAT': 'IMG',
                'UNIT': '',
                'FUNCTION': 'get_windrose',
                'PATH': 'path_windrose',
                'VARIABLES': ['path_windrose'],
                'LIEN UTILE': 'https://meteo.data.gouv.fr/',
            },
            {
                'THEMATIQUE': 'Localisation du site',
                'INFORMATION': 'Altitude',
                'SOURCE': 'https://geoservices.ign.fr/bdalti',
                'DESCRIPTION': 'Altitude',
                'FORMAT': 'FLOAT',
                'UNIT': 'm',
                'FUNCTION': 'get_altitude',
                'VARIABLES': ['altitude'],
                'PATH': 'path_altitude',
                'LIEN UTILE': 'https://geoservices.ign.fr/bdalti',
            },
            {
                'THEMATIQUE': 'Risque naturel',
                'INFORMATION': 'Inondation',
                'SOURCE': 'https://www.georisques.gouv.fr/donnees/bases-de-donnees/zonages-inondation-rapportage-2020',
                'DESCRIPTION': 'Information géorisque',
                'FORMAT': 'STR',
                'UNIT': '',
                'FUNCTION': 'get_zonage_inondation',
                'VARIABLES': ['info_zonage_inondations'],
                'PATH': 'path_zonage_inondations',
                'LIEN UTILE': 'https://www.georisques.gouv.fr/donnees/bases-de-donnees/zonages-inondation-rapportage-2020',
            },
            {
                'THEMATIQUE': 'Risque naturel',
                'INFORMATION': 'Mouvement de terrain',
                'SOURCE': 'https://www.georisques.gouv.fr',
                'DESCRIPTION': 'Information géorisque',
                'FORMAT': 'STR',
                'UNIT': '',
                'FUNCTION': 'get_mouvement_de_terrain',
                'VARIABLES': ['path_mvt_terrain'],
                'PATH': 'path_mvt_terrain',
                'LIEN UTILE': 'https://www.georisques.gouv.fr/risques/mouvements-de-terrain/donnees#/',
            },
            {
                'THEMATIQUE': 'Risque naturel',
                'INFORMATION': 'Aléas retrait-gonflement des argiles',
                'SOURCE': 'https://www.georisques.gouv.fr/donnees/bases-de-donnees/retrait-gonflement-des-argiles',
                'DESCRIPTION': 'Information géorisque',
                'FORMAT': 'STR',
                'UNIT': '',
                'FUNCTION': 'get_retrait_argile',
                'VARIABLES': ['path_retrait_argile'],
                'PATH': 'path_retrait_argile',
                'LIEN UTILE': 'https://www.georisques.gouv.fr/donnees/bases-de-donnees/retrait-gonflement-des-argiles',
            },
            # {"THEMATIQUE": "Localisation du site", "INFORMATION": "Le projet est dans une zone déjà construite",
            #  "SOURCE": "https://geoservices.ign.fr/documentation/donnees/vecteur/bdtopo", "DESCRIPTION": "oui/non",
            #  "FORMAT": "STR", "UNIT": "", "FUNCTION": "get_localisation_du_site",
            #  "VARIABLES": ["path_retrait_argile"], "LIEN UTILE": "https://geoservices.ign.fr/bdtopo"},
            {
                'THEMATIQUE': 'Localisation du site',
                'INFORMATION': 'Rues délimitantes',
                'SOURCE': 'https://wiki.openstreetmap.org/wiki/API',
                'DESCRIPTION': 'Noms des rues',
                'FORMAT': list[str],
                'FUNCTION': 'get_rues_delimitantes',
                'VARIABLES': ['rues_delimitantes'],
                'PATH': 'path_rues_delimitantes',
                'LIEN UTILE': 'https://www.openstreetmap.org/#map=5/46.449/2.210',
            },
            {
                'THEMATIQUE': 'Contexte hydrographique',
                'INFORMATION': "Cours d'eau qui traverse la ville",
                'SOURCE': 'https://www.sandre.eaufrance.fr/atlas/srv/fre/catalog.search#/metadata/3b3d3c56-d9b6-4625-a57e-ba054e798274',
                'DESCRIPTION': "Noms des cours d'eau",
                'FORMAT': 'STR',
                'FUNCTION': 'get_liste_cours_deau',
                'VARIABLES': ['path_liste_cours_deau'],
                'PATH': 'path_liste_cours_deau',
                'LIEN UTILE': 'https://geocatalogue.apur.org/catalogue/srv/api/records/7fa4c224-fe38-4e2c-846d-dcc2fa7ef73e',
            },
            {
                'THEMATIQUE': 'Contexte hydrographique',
                'INFORMATION': 'Bassin versant',
                'SOURCE': 'https://www.sandre.eaufrance.fr/atlas/srv/fre/catalog.search#/metadata/3b3d3c56-d9b6-4625-a57e-ba054e798274',
                'DESCRIPTION': 'Carte avec mention du bassin versant',
                'FORMAT': 'STR',
                'FUNCTION': 'get_bassin_versant',
                'VARIABLES': ['path_bassin_versant'],
                'PATH': 'path_bassin_versant',
                'LIEN UTILE': 'https://geocatalogue.apur.org/catalogue/srv/api/records/7fa4c224-fe38-4e2c-846d-dcc2fa7ef73e',
            },
            {
                'THEMATIQUE': 'Risque naturel',
                'INFORMATION': 'Remontée de nappe',
                'SOURCE': 'https://www.georisques.gouv.fr/donnees/bases-de-donnees/inondations-par-remontee-de-nappes',
                'DESCRIPTION': 'oui / non et noms des zones',
                'FORMAT': 'STR',
                'FUNCTION': 'get_remontee_nappe',
                'VARIABLES': ['path_remontee_nappe'],
                'PATH': 'path_remontee_nappe',
                'LIEN UTILE': 'https://www.georisques.gouv.fr/donnees/bases-de-donnees/inondations-par-remontee-de-nappes',
            },
            {
                'THEMATIQUE': 'Risque naturel',
                'INFORMATION': 'Sismique',
                'SOURCE': 'https://www.georisques.gouv.fr/doc-api#!/Zonage32Sismique/rechercheRadonUsingGET_1',
                'DESCRIPTION': 'Information géorisque',
                'FORMAT': 'STR',
                'FUNCTION': 'get_zone_sismique',
                'VARIABLES': ['path_zonage_sismique'],
                'PATH': 'path_zonage_sismique',
                'LIEN UTILE': 'https://www.data.gouv.fr/fr/datasets/zonage-sismique-de-la-france-1/',
            },
            {
                'THEMATIQUE': 'Risque naturel',
                'INFORMATION': 'Radon',
                'SOURCE': 'https://www.georisques.gouv.fr/doc-api#/Radon',
                'DESCRIPTION': 'oui / non',
                'FORMAT': 'STR',
                'FUNCTION': 'get_radon',
                'VARIABLES': ['path_radon'],
                'PATH': 'path_radon',
                'LIEN UTILE': 'https://www.irsn.fr/savoir-comprendre/environnement/connaitre-potentiel-radon-ma-commune',
            },
            # {
            #     "THEMATIQUE": "Réseaux",
            #     "INFORMATION": "Présence des réseaux: électricité, gaz, réseau de chaleur",
            #     "SOURCE": "https://www.data.gouv.fr/fr/reuses/cartographie-des-reseaux-electriques-et-gaziers/",
            #     "DESCRIPTION": "oui / non",
            #     "FORMAT": "STR",
            #     "FUNCTION": self.get_presence_reseaux(),
            #     "VARIABLES": self.path_radon
            # },
            # {
            #     "THEMATIQUE": "Réseaux",
            #     "INFORMATION": "Présence des réseaux: électricité, gaz, réseau de chaleur",
            #     "SOURCE": "https://data.enedis.fr/api/explore/v2.1/console",
            #     "DESCRIPTION": "",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Réseaux",
            #     "INFORMATION": "Présence des réseaux: électricité, gaz, réseau de chaleur",
            #     "SOURCE": "https://data.enedis.fr/pages/cartographie-des-reseaux-contenu/",
            #     "DESCRIPTION": "",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Risque technologique",
            #     "INFORMATION": "Sites industriels",
            #     "SOURCE": "https://www.georisques.gouv.fr/donnees/bases-de-donnees/inventaire-historique-de-sites-industriels-et-activites-de-service",
            #     "DESCRIPTION": "Information géorisque",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Environnement proche",
            #     "INFORMATION": "Hauteurs de bâtiments alentour et ombres portées",
            #     "SOURCE": "https://github.com/bbrangeo/pymdu",
            #     "DESCRIPTION": "Carte avec % d'ombres portées sur la zone",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            {
                'THEMATIQUE': 'Climatologie',
                'INFORMATION': "Durée d'ensoleillement et course du soleil",
                'SOURCE': 'https://pvlib-python.readthedocs.io/en/v0.9.0/auto_examples/plot_sunpath_diagrams.html',
                'DESCRIPTION': 'Héliodon',
                'FORMAT': 'IMG',
                'FUNCTION': 'get_course_du_soleil',
                'VARIABLES': ['path_course_soleil'],
                'PATH': 'path_course_soleil',
                'LIEN UTILE': 'https://www.sunearthtools.com/dp/tools/pos_sun.php?lang=fr&utc=1&point=48.8583',
            },  # {
            #     "THEMATIQUE": "Risque naturel",
            #     "INFORMATION": "Termites et insectes xylophages",
            #     "SOURCE": "https://termite.com.fr/rechercher/resultats-de-votre-recherche",
            #     "DESCRIPTION": "Niveau d'infestation",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Contexte géologique",
            #     "INFORMATION": "Description du profil géologique du sol",
            #     "SOURCE": "https://entrepot.recherche.data.gouv.fr/dataset.xhtml?persistentId=doi:10.15454/BPN57S",
            #     "DESCRIPTION": "Schéma de la couche de sol",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Nuisance acoustique",
            #     "INFORMATION": "Infrastructure proche à moins de 500 m",
            #     "SOURCE": "https://www.ecologie.gouv.fr/cartes-strategiques-bruit-csb-et-plans-prevention-du-bruit-dans-lenvironnement-ppbe-autour-des",
            #     "DESCRIPTION": "Liste",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Nuisance olfactive",
            #     "INFORMATION": "Infrastructure proche à moins de 500 m",
            #     "SOURCE": "XXX",
            #     "DESCRIPTION": "Liste",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Nuisance électromagnétique",
            #     "INFORMATION": "Infrastructure proche à moins de 500 m",
            #     "SOURCE": "https://www.cartoradio.fr/#/cartographie/lonlat/-1.171305/46.158141",
            #     "DESCRIPTION": "Liste",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Localisation du site",
            #     "INFORMATION": "Description des zones naturelles éventuellement protégées",
            #     "SOURCE": "https://www.ecologie.gouv.fr/aires-protegees-en-france#:~:text=En%20mai%202022%2C%20355%20r%C3%A9serves,7%20r%C3%A9serves%20naturelles%20de%20Corse.",
            #     "DESCRIPTION": "oui / non et noms des zones",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Accessibilité du site",
            #     "INFORMATION": "Localisation et fréquence des arrêts de bus. Le transport en commun sont à moins de 10 min à pied ou à 600 m du site",
            #     "SOURCE": "https://wiki.openstreetmap.org/wiki/API",
            #     "DESCRIPTION": "Carte et mention oui/non",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Accessibilité du site",
            #     "INFORMATION": "Présence de cheminement piéton",
            #     "SOURCE": "https://wiki.openstreetmap.org/wiki/API",
            #     "DESCRIPTION": "Carte et mention oui/non",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            {
                'THEMATIQUE': 'Services de proximité',
                'INFORMATION': 'Liste des commerces et services présents à moins de 10 min à pieds (boulangerie, charcuterie, épicerie, supermarché…).\nPoste, mairie, banque, assurance…',
                'SOURCE': 'https://wiki.openstreetmap.org/wiki/API',
                'DESCRIPTION': 'Carte et mention oui/non + liste',
                'FORMAT': 'IMG',
                'FUNCTION': 'get_map_commerces',
                'VARIABLES': ['path_map_commerces'],
                'PATH': 'path_map_commerces',
                'LIEN UTILE': 'https://www.openstreetmap.org/',
            },
            {
                'THEMATIQUE': 'Services de proximité',
                'INFORMATION': 'Des équipements culturels et/ou de loisirs sont à moinds de 10 min à pied du site ou sont crées dans le cadre du projet.\nCinémas, bibliothèque, musée, théâtre, librairie…',
                'SOURCE': 'https://wiki.openstreetmap.org/wiki/API',
                'DESCRIPTION': 'Carte et mention oui/non + liste',
                'FORMAT': 'IMG',
                'FUNCTION': 'get_map_equipements_culturels_loisirs',
                'VARIABLES': ['path_map_equipements_culturels_loisirs'],
                'PATH': 'path_map_equipements_culturels_loisirs',
                'LIEN UTILE': 'https://www.openstreetmap.org/',
            },  #
            {
                'THEMATIQUE': 'Services de proximité',
                'INFORMATION': 'Des services de santé sont à moins de 10 min à pied du site',
                'SOURCE': 'https://wiki.openstreetmap.org/wiki/API',
                'DESCRIPTION': 'Carte et mention oui/non + liste',
                'FORMAT': 'IMG',
                'FUNCTION': 'get_map_service_sante',
                'VARIABLES': ['path_map_service_sante'],
                'PATH': 'path_map_service_sante',
                'LIEN UTILE': 'https://www.openstreetmap.org/',
            },
            {
                'THEMATIQUE': 'Services de proximité',
                'INFORMATION': 'Des établissements scolaires sont à moins de 10 min à pied du site',
                'SOURCE': 'https://wiki.openstreetmap.org/wiki/API',
                'DESCRIPTION': 'Carte et mention oui/non + liste',
                'FORMAT': 'IMG',
                'FUNCTION': 'get_map_etablissements_scolaires',
                'VARIABLES': ['path_map_etablissements_scolaires'],
                'PATH': 'path_map_etablissements_scolaires',
                'LIEN UTILE': 'https://www.openstreetmap.org/',
            },  # {
            #     "THEMATIQUE": "Réseaux",
            #     "INFORMATION": "Eau, assainissement, télécom",
            #     "SOURCE": "https://www.reseaux-et-canalisations.ineris.fr/gu-presentation/construire-sans-detruire/teleservice-reseaux-et-canalisations.html",
            #     "DESCRIPTION": "oui / non",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Ressources",
            #     "INFORMATION": "Ressources géothermiques",
            #     "SOURCE": "https://www.geothermies.fr/viewer/?extent=-455664.6447%2C5414187.5875%2C472586.6268%2C5842234.9459&al=region/NOA",
            #     "DESCRIPTION": "Carte",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Accessibilité du site",
            #     "INFORMATION": "Présence de station d'auto-partage à moins de 10 min à pied du site",
            #     "SOURCE": "https://wiki.openstreetmap.org/wiki/API",
            #     "DESCRIPTION": "Carte et mention oui/non",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Accessibilité du site",
            #     "INFORMATION": "Lignes desservant le site: bus, tramway, train…",
            #     "SOURCE": "https://wiki.openstreetmap.org/wiki/API",
            #     "DESCRIPTION": "Carte",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # },
            # {
            #     "THEMATIQUE": "Accessibilité du site",
            #     "INFORMATION": "Présence de pistes cyclables",
            #     "SOURCE": "https://wiki.openstreetmap.org/wiki/API",
            #     "DESCRIPTION": "Carte et mention oui/non",
            #     "FUNCTION": "",
            #     "VARIABLES": ""
            # }
        ]

        with Parallel(verbose=100, n_jobs=10) as parallel:
            delayed_funcs = [
                delayed(self.worker)(run, self) for run in configuration_diagnostic
            ]
            parallel_pool = parallel(delayed_funcs)

        df = pd.DataFrame.from_dict(parallel_pool)
        # Create a Pandas Excel writer using XlsxWriter as the engine.
        writer = pd.ExcelWriter(
            path=os.path.join(self.output_path, filename), engine='xlsxwriter'
        )
        # Close the Pandas Excel writer and output the Excel file.
        df.to_excel(
            writer, sheet_name='Diagnostic', startrow=1, header=False, index=False
        )
        # Get the xlsxwriter workbook and worksheet objects.
        workbook = writer.book
        worksheet = writer.sheets['Diagnostic']
        (max_row, max_col) = df.shape
        column_settings = [{'header': column} for column in df.columns]

        worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})
        # Autofit the worksheet.
        worksheet.autofit()
        path_col_index = df.columns.get_loc('PATH')

        for i, row in df.iterrows():
            worksheet.write_url(row=i + 1, col=path_col_index, url=row['PATH'])
            if 'IMG' == row['FORMAT']:
                worksheet = workbook.add_worksheet(name=f"{row['FUNCTION'][:31]}")
                # worksheet.embed_image(f"J{i + 2}", filename=f"{row['PATHS'][0]['path']}",
                #                       options={"url": f"{row['PATHS'][0]['path']}", "x_scale": 1.5, "y_scale": 1.5}),
                (
                    worksheet.insert_image(
                        'B1',
                        filename=f"{row['PATH']}",
                        options={
                            'url': f"{row['PATH']}",
                            'x_scale': 0.5,
                            'y_scale': 0.5,
                        },
                    ),
                )
        # Set the background image.
        # worksheet.set_background("logo.png")

        # Autofit the worksheet.
        worksheet.autofit()
        arial_format = workbook.add_format({'font_name': 'Arial'})
        worksheet.set_column(0, len(df.columns) - 1, None, arial_format)

        # # # Add a header format.
        # header_format = workbook.add_format({
        #     'bold': True,
        #     'text_wrap': True,
        #     'valign': 'top',
        #     'fg_color': '#D7E4BC',
        #     'border': 1})
        #
        # # Write the column headers with the defined format.
        # for col_num, value in enumerate(df.columns.values):
        #     worksheet.write(0, col_num + 1, value, header_format)
        writer.close()

        self.modifier_feuille_diagnostic(filename)

    def modifier_feuille_diagnostic(self, filename):
        chemin_fichier_excel = os.path.join(self.output_path, filename)
        # Charger le classeur Excel
        wb = load_workbook(chemin_fichier_excel)

        # Sélectionner la feuille "Diagnostic"
        ws = wb['Diagnostic']

        # Définir la police Arial
        arial_font = Font(name='Arial')

        # Supprimer les couleurs de fond des cellules et appliquer la police Arial
        for row in ws.iter_rows():
            for cell in row:
                cell.fill = PatternFill(
                    fill_type=None
                )  # Suppression de la couleur de fond
                cell.font = arial_font

        # Créer un nouveau nom de fichier avec "_miseEnforme" à la fin
        nouveau_nom_fichier = chemin_fichier_excel.replace('.xlsx', '_miseEnForme.xlsx')

        # Enregistrer les modifications dans le nouveau classeur Excel
        wb.save(nouveau_nom_fichier)
        ws.conditional_formatting = []

        # Retourner le nouveau nom de fichier pour référence
        return nouveau_nom_fichier

    @staticmethod
    def worker(run, class_diag):
        fun = run['FUNCTION']
        func = getattr(class_diag, fun)
        func()
        run['RESULTS'] = []
        for var in run['VARIABLES']:
            result = getattr(class_diag, var)
            run['RESULTS'].append(result)
            if isinstance(run['PATH'], str) and 'path' in run['PATH']:
                path = getattr(class_diag, run['PATH'])
                run['PATH'] = path
        return run

    def get_localisation_du_site(self, filename='LocalisationDuSite.json') -> tuple:
        self.path_localisation_du_site = os.path.join(self.output_path, filename)

        params = {'q': self.address, 'type': 'housenumber', 'autocomplete': 1}
        r = requests.get('https://api-adresse.data.gouv.fr/search/', params=params)
        adresse = r.json()
        with open(self.path_localisation_du_site, 'w') as f:
            json.dump(adresse, f)

        self.lon = adresse['features'][0]['geometry']['coordinates'][0]
        self.lat = adresse['features'][0]['geometry']['coordinates'][1]
        return (self.lon, self.lat)

    def get_parcelle_cadastrale(self, filename='ParcelleCadastrale.json') -> float:
        self.path_parcelle_cadastrale = os.path.join(self.output_path, filename)

        self.lon, self.lat = self.get_localisation_du_site()
        self.point = Point(self.lon, self.lat)
        geojson = json.loads(shapely.to_geojson(self.point))

        params = {'geom': json.dumps(geojson), 'source_ign': 'PCI'}
        r = requests.get(
            'https://apicarto.ign.fr/api/cadastre/parcelle',
            params=params,
            headers={'accept': 'application/json'},
        )
        parcelle = r.json()
        with open(self.path_parcelle_cadastrale, 'w') as f:
            json.dump(parcelle, f)

        self.parcelle_number = parcelle['features'][0]['properties']['numero']
        self.parcelle_commnune = parcelle['features'][0]['properties']['nom_com']
        self.parcelle_surface = parcelle['features'][0]['properties']['contenance']
        self.insee_code = parcelle['features'][0]['properties']['code_insee']
        self.department = parcelle['features'][0]['properties']['code_dep']

        gdf = gpd.GeoDataFrame.from_features(parcelle['features'])
        self.parcelle_bbox = [
            gdf.bounds.minx,
            gdf.bounds.miny,
            gdf.bounds.maxx,
            gdf.bounds.maxy,
        ]

        return self.parcelle_number

    def get_zonage_plu(self, filename='ZonagePlu.json') -> str:
        self.path_zonage_plu = os.path.join(self.output_path, filename)

        self.lon, self.lat = self.get_localisation_du_site()
        self.point = Point(self.lon, self.lat)
        geojson = json.loads(shapely.to_geojson(self.point))

        params = {'geom': geojson}
        r = requests.get('https://apicarto.ign.fr/api/gpu/zone-urba', json=params)
        plu = r.json()
        with open(self.path_zonage_plu, 'w') as f:
            json.dump(plu, f)

        self.zonage_plu = plu['features'][0]['properties']['libelle']

        return self.zonage_plu

    def get_altitude(self, filename: str = 'Altitude.json') -> float:
        self.path_altitude = os.path.join(self.output_path, filename)

        self.lon, self.lat = self.get_localisation_du_site()
        params = {'lon': self.lon, 'lat': self.lat, 'zonly': 'true'}
        r = requests.get(
            'https://wxs.ign.fr/calcul/alti/rest/elevation.json', params=params
        )
        altitude = r.json()
        with open(self.path_altitude, 'w') as f:
            json.dump(altitude, f)
        self.altitude = altitude['elevations'][0]
        return self.altitude

    def get_streets_in_bbox(self, filename: str = 'RuesDelimitantes.json') -> list:
        self.path_rues_delimitantes = os.path.join(self.output_path, filename)
        if self.parcelle_bbox is None:
            self.get_parcelle_cadastrale()
        min_lat, min_lon, max_lat, max_lon = (
            self.parcelle_bbox[0],
            self.parcelle_bbox[1],
            self.parcelle_bbox[2],
            self.parcelle_bbox[3],
        )
        url = f'https://nominatim.openstreetmap.org/search?format=json&q=&bounded=1&viewbox={min_lon},{min_lat},{max_lon},{max_lat}'
        response = requests.get(url)
        data = response.json()
        with open(self.path_rues_delimitantes, 'w') as f:
            json.dump(data, f)

        street_names = []
        for item in data:
            if 'address' in item:
                street_name = item['address'].get('road', 'Unknown')
                street_names.append(street_name)

        return street_names

    def get_rues_delimitantes(self) -> list:
        self.rues_delimitantes = self.get_streets_in_bbox()
        return self.rues_delimitantes

    def find_url_meteo(self):
        if self.department is None:
            self.get_parcelle_cadastrale()

        response = requests.get(
            f'https://www.data.gouv.fr/api/2/datasets/6569b4473bedf2e7abad3b72/resources//?page=1&page_size=30&type=main&q=HOR_departement_{self.department}_periode_'
        )

        data = response.json()

        df_json = pd.DataFrame.from_dict(data['data'])

        # ======
        init = []
        end = []
        for x in list(df_json.title):
            period = x.split('_')[-1]
            init.append(float(period.split('-')[0]))
            end.append(float(period.split('-')[1]))
        df_json['init'] = init
        df_json['end'] = end
        # ======

        for i, e, url in zip(df_json.init, df_json.end, df_json.url):
            if (self.year > i) & (self.year < e):
                return url

    def get_meteo_locale(self):
        url = self.find_url_meteo()
        data = pd.read_csv(url, compression='gzip', header=0, sep=';', quotechar='"')
        from datetime import datetime

        data['date'] = [
            datetime.strptime(str(x), '%Y%m%d%H').strftime('%Y-%m-%d %H:00:00')
            for x in data['AAAAMMJJHH']
        ]

        all_stations = {}
        for station in list(dict.fromkeys(data['NOM_USUEL'].values)):
            all_stations[station] = {
                'type': 'Point',
                'LAT': data[data.NOM_USUEL == station].reset_index().LAT[0],
                'LON': data[data.NOM_USUEL == station].reset_index().LON[0],
                'coordinates': [
                    data[data.NOM_USUEL == station].reset_index().LAT[0],
                    data[data.NOM_USUEL == station].reset_index().LON[0],
                ],
            }

        self.StationsDuDepartement = pd.DataFrame.from_dict(all_stations).T
        df = self.StationsDuDepartement.reset_index()

        # Find the nearest point
        nearest_point, nearest_index = self.__find_nearest_point()

        self.StationLaPlusProche = df[df.index == nearest_index]

        data = data[data.NOM_USUEL == self.StationLaPlusProche['index'].values[0]]
        df = data[
            [
                'date',
                ' T',
                'FF2',
                'DD2',
                ' FXI',
                'DXI',
                'LAT',
                'LON',
                'NOM_USUEL',
                'RR1',
            ]
        ]
        df.index = pd.to_datetime(df.date)
        df.columns = [x.strip() for x in df.columns]

        self.weather_data = df

    def get_temperature_moyenne_mensuelle(
        self, filename='Temperature_moyenne_mensuelle.png'
    ) -> str:
        if self.weather_data is None:
            self.get_meteo_locale()
        self.weather_data.index = pd.to_datetime(self.weather_data.date)
        self.temperature_moyenne_mensuelle = self.weather_data['T'].resample('M').mean()
        sns.set_context('talk')
        fig, (ax) = plt.subplots(1, 1, figsize=(15, 5))

        ax.plot(
            self.weather_data['T'].resample('M').mean().index,
            self.weather_data['T'].resample('M').mean(),
            color='k',
            linewidth='2',
        )
        ax.plot(
            self.weather_data['T'].resample('M').max().index,
            self.weather_data['T'].resample('M').max(),
            color='k',
            linewidth='0.5',
            linestyle='--',
        )
        ax.plot(
            self.weather_data['T'].resample('M').min().index,
            self.weather_data['T'].resample('M').min(),
            color='k',
            linewidth='0.5',
            linestyle='--',
        )

        ax.fill_between(
            self.weather_data['T'].resample('M').max().index,
            self.weather_data['T'].resample('M').min(),
            self.weather_data['T'].resample('M').max(),
            alpha=0.2,
            color='grey',
            hatch='//',
        )
        ax.set_title(self.StationLaPlusProche['index'].values[0])
        ax.set_ylabel('Temp \n ext. C')
        self.path_temperature_moyenne_mensuelle = os.path.join(
            self.output_path, filename
        )
        plt.savefig(self.path_temperature_moyenne_mensuelle)
        return self.path_temperature_moyenne_mensuelle

    def get_pluviometrie_moyenne_mensuelle(
        self, filename='Pluviometrie_moyenne_mensuelle.png'
    ) -> str:
        if self.weather_data is None:
            self.get_meteo_locale()
        self.weather_data.index = pd.to_datetime(self.weather_data.date)
        self.pluviometrie_moyenne_mensuelle = (
            self.weather_data['RR1'].resample('M').mean()
        )
        sns.set_context('talk')

        fig, (ax) = plt.subplots(1, 1, figsize=(15, 5))

        self.weather_data['RR1'].resample('M').mean().plot(
            ax=ax, color='blue', linewidth=2
        )
        ax.set_title(self.StationLaPlusProche['index'].values[0])
        ax.set_ylabel("Pluie dans \n l'heure (en mm)")
        self.path_pluviometrie_moyenne_mensuelle = os.path.join(
            self.output_path, filename
        )
        plt.savefig(self.path_pluviometrie_moyenne_mensuelle)
        return self.path_pluviometrie_moyenne_mensuelle

    def get_windrose(self, filename='Windrose.png') -> str:
        if self.weather_data is None:
            self.get_meteo_locale()
        self.path_windrose = self.output_path + filename

        df = pd.DataFrame()
        df.index = pd.to_datetime(self.weather_data.date)
        df['year'] = self.weather_data.index.year
        df['speed'] = self.weather_data['FXI']
        df['direction'] = self.weather_data['DXI']

        ax = WindroseAxes.from_ax()
        ax.bar(
            df['direction'].values,
            df['speed'].values,
            bins=np.arange(0.01, 30, 4),
            cmap=cm.coolwarm_r,
            normed=True,
            lw=3,
        )

        plt.legend(
            title='Vitesse du vent [km/h]',
            fontsize='15',
            bbox_to_anchor=(1.1, 0.3),
            fancybox=True,
        )
        plt.title(
            self.StationLaPlusProche['index'].values[0]
            + '\n'
            + str(list(dict.fromkeys(df.year.values)))
            + '\n- force maximale du vent instantané dans l’heure, mesurée à 10 m (en m/s et '
            '1/10)\n' + '- direction de FXI (en rose de 360)'
        )
        self.path_windrose = os.path.join(self.output_path, filename)
        plt.savefig(self.path_windrose)
        return self.path_windrose

    def get_course_du_soleil(self, filename='CourseDuSoleil.png') -> str:
        self.path_course_soleil = os.path.join(self.output_path, filename)
        if self.lon is None or self.lat is None:
            self.get_localisation_du_site()

        SunPath = pd.DataFrame()
        thetime = datetime(2018, 5, 23, 13)
        lat = self.lat
        lon = self.lon
        tz = 1
        liste_alt = []
        liste_azm = []
        times = pd.date_range(start='2024-01-01', end='2024-12-31', freq='1h')
        for time in times:
            thetime = time.to_pydatetime()
            alt, azm = sunpos(thetime, lat, lon, tz, dst=False)
            liste_alt.append(alt)
            liste_azm.append(azm)
        SunPath['alt'] = liste_alt
        SunPath['azm'] = liste_azm
        SunPath['index'] = pd.to_datetime(times)
        SunPath.index = pd.to_datetime(times)
        SunPath['hour'] = SunPath.index.hour

        sns.set_context('talk')
        fig, (ax) = plt.subplots(1, 1, figsize=(15, 10))
        SunPath = SunPath[(SunPath.hour > 6) & (SunPath.hour < 22)]
        sns.scatterplot(
            data=SunPath,
            x='azm',
            y='alt',
            s=5,
            palette='coolwarm',
            hue='hour',
            legend='full',
        )
        ax.set_ylim([0, 70])

        # ax.legend(bbox_to_anchor=(1, 1.2))
        ax.legend()
        plt.savefig(self.path_course_soleil)
        return self.path_course_soleil

    def get_radon(self, filename='Radon.json') -> str:
        self.path_radon = os.path.join(self.output_path, filename)
        if self.insee_code is None:
            self.get_parcelle_cadastrale()
        params = {'code_insee': self.insee_code}

        r = requests.get('https://georisques.gouv.fr/api/v1/radon', params=params)
        radon = r.json()
        with open(self.path_radon, 'w') as f:
            json.dump(radon, f)

        return self.path_radon

    def get_zonage_inondation(self, filename='ZonageInondation.json') -> tuple:
        self.path_zonage_inondations = os.path.join(self.output_path, filename)
        if self.lon is None or self.lat is None:
            self.get_localisation_du_site()
        params = {'latlon': f'{self.lon},{self.lat}', 'rayon': 1000, 'page': 1}
        r = requests.get('https://georisques.gouv.fr/api/v1/gaspar/azi', params=params)
        zonage_inondations = r.json()
        with open(self.path_zonage_inondations, 'w') as f:
            json.dump(zonage_inondations, f)

        self.info_zonage_inondations = zonage_inondations['data']

        return self.path_zonage_inondations, self.info_zonage_inondations

    def get_mouvement_de_terrain(self, filename='MouvementDeTerrain.json') -> str:
        if self.lon is None or self.lat is None:
            self.get_localisation_du_site()
        self.path_mvt_terrain = os.path.join(self.output_path, filename)
        params = {'latlon': f'{self.lon},{self.lat}', 'page': 1}
        r = requests.get('https://georisques.gouv.fr/api/v1/mvt', params=params)
        mvtTerrain = r.json()
        with open(self.path_mvt_terrain, 'w') as f:
            json.dump(mvtTerrain, f)
        return self.path_mvt_terrain

    def get_retrait_argile(self, filename='RetraitArgile.json') -> str:
        if self.lon is None or self.lat is None:
            self.get_localisation_du_site()

        self.path_retrait_argile = os.path.join(self.output_path, filename)
        params = {'latlon': f'{self.lon},{self.lat}', 'page': 1}
        r = requests.get('https://georisques.gouv.fr/api/v1/rga', params=params)
        rga = r.json()
        with open(self.path_retrait_argile, 'w') as f:
            json.dump(rga, f)
        return self.path_retrait_argile

    def get_remontee_nappe(self, filename='RemonteeNappe.zip') -> str:
        if self.lon is None or self.lat is None:
            self.get_parcelle_cadastrale()
        self.path_remontee_nappe = os.path.join(self.output_path, filename)
        remontee_nappe = requests.get(
            f'https://files.georisques.fr/REMNAPPES/Dept_{self.department}.zip'
        )
        with open(self.path_remontee_nappe, 'wb') as fd:
            for chunk in remontee_nappe.iter_content(chunk_size=128):
                fd.write(chunk)

        return self.path_remontee_nappe

    def get_zone_sismique(self, filename='ZoneSismique.json') -> str:
        if self.lon is None or self.lat is None:
            self.get_localisation_du_site()
        self.path_zonage_sismique = os.path.join(self.output_path, filename)
        params = {'latlon': f'{self.lon},{self.lat}', 'page': 1}
        r = requests.get(
            'https://georisques.gouv.fr/api/v1/zonage_sismique', params=params
        )
        zonage_sismique = r.json()
        with open(self.path_zonage_sismique, 'w') as f:
            json.dump(zonage_sismique, f)
        return self.path_zonage_sismique

    def from_point_to_bbox(self, meters=1000):
        if self.lon is None or self.lat is None:
            self.get_localisation_du_site()
        bearings = [0, 90]
        origin = geopy.Point(self.lat, self.lon)
        l = []

        for bearing in bearings:
            destination = geodesic(meters=meters).destination(origin, bearing)
            coords = destination.longitude, destination.latitude
            l.extend(coords)
        # xmin, ymin, xmax, ymax
        return l

    def get_bassin_versant(self, filename='Bassin_versant.json') -> str:
        self.path_bassin_versant = os.path.join(self.output_path, filename)
        list_bbox = self.from_point_to_bbox()
        reverse_list = [list_bbox[0], list_bbox[3], list_bbox[2], list_bbox[1]]

        params = {
            'SERVICE': 'WFS',
            'TYPENAME': 'BDTOPO_V3:bassin_versant_topographique',
            'VERSION': '2.0.0',
            'REQUEST': 'GetFeature',
            'SRSNAME': 'EPSG:4326',
            'BBOX': f'{reverse_list[0]}, {reverse_list[1]}, {reverse_list[2]}, {reverse_list[3]}, EPSG:4326',
            'EPSG': 4326,
            'STARTINDEX': 0,
            'COUNT': 1000,
            'outputFormat': 'application/json',
        }
        bassin_versant = requests.get(
            'https://data.geopf.fr/wfs/ows', params=params
        ).json()

        with open(self.path_bassin_versant, 'w') as f:
            json.dump(bassin_versant, f)
        return self.path_bassin_versant

    def get_liste_cours_deau(self, filename='Liste_cours_deau.json'):
        self.path_liste_cours_deau = os.path.join(self.output_path, filename)
        list_bbox = self.from_point_to_bbox()
        reverse_list = [list_bbox[0], list_bbox[3], list_bbox[2], list_bbox[1]]
        params = {
            'SERVICE': 'WFS',
            'TYPENAME': 'BDTOPO_V3:cours_d_eau',
            'VERSION': '2.0.0',
            'REQUEST': 'GetFeature',
            'SRSNAME': 'EPSG:4326',
            'BBOX': f'{reverse_list[0]}, {reverse_list[1]}, {reverse_list[2]}, {reverse_list[3]}, EPSG:4326',
            'EPSG': 4326,
            'STARTINDEX': 0,
            'COUNT': 1000,
            'outputFormat': 'application/json',
        }
        cours_d_eau = requests.get(
            'https://data.geopf.fr/wfs/ows', params=params
        ).json()
        with open(self.path_liste_cours_deau, 'w') as f:
            json.dump(cours_d_eau, f)
        return self.path_liste_cours_deau

    def get_isochrones(self, filename='isochrones.json'):
        if self.lon is None or self.lat is None:
            self.get_localisation_du_site()
        self.path_isochrones = os.path.join(self.output_path, filename)
        params = {
            'resource': 'bdtopo-pgr',
            'profile': 'pedestrian',
            'costType': 'distance',
            'costValue': '1000',
            'direction': 'departure',
            'point': f'{self.lon},{self.lat}',
            'geometryFormat': 'geojson',
        }
        isochrones = requests.get(
            url='https://wxs.ign.fr/calcul/geoportail/isochrone/rest/1.0.0/isochrone',
            params=params,
        ).json()

        with open(self.path_isochrones, 'w') as f:
            json.dump(isochrones, f)
        return self.path_isochrones

    def get_map_service_sante(self) -> dict:
        if self.lon is None or self.lat is None:
            self.get_localisation_du_site()
        # Créer une map centrée sur les coordonnées spécifiées
        map = folium.Map(location=[self.lat, self.lon], zoom_start=13)

        # Définir le rayon de recherche en mètres (10000 mètres dans cet exemple)
        rayon = 10000

        # Requête Overpass API pour récupérer les données des établissements de santé
        overpass_url = 'http://overpass-api.de/api/interpreter'
        overpass_query = """
                [out:json];
                (
                  node["amenity"="hospital"](around:{}, {}, {});
                  node["amenity"="clinic"](around:{}, {}, {});
                  node["amenity"="doctors"](around:{}, {}, {});
                );
                out body;
                >;
                out skel qt;
                """.format(
            rayon,
            self.lat,
            self.lon,
            rayon,
            self.lat,
            self.lon,
            rayon,
            self.lat,
            self.lon,
        )

        response = requests.get(overpass_url, params={'data': overpass_query})
        data = response.json()

        # Parcourir les données et ajouter des marqueurs sur la map pour chaque établissement de santé trouvé
        for element in data['elements']:
            if 'tags' in element:
                if 'name' in element['tags']:
                    nom = element['tags']['name']
                else:
                    nom = 'Service de santé'
                lat = element['lat']
                lon = element['lon']
                folium.Marker(
                    [lat, lon], popup=nom, icon=folium.Icon(color='red', icon='plus')
                ).add_to(map)

        # Afficher la map dans le notebook
        folium.Marker([self.lat, self.lon], popup='Votre emplacement').add_to(
            map
        )  # display(map)
        # Sauvegarder la map dans un fichier PNG
        self.path_map_service_sante = self.__convert_map_png(
            folium_map=map, file_name='service_sante'
        )
        return data

    def get_map_etablissements_scolaires(self) -> dict:
        if self.lon is None or self.lat is None:
            self.get_localisation_du_site()

        # Créer une map centrée sur les coordonnées spécifiées
        map = folium.Map(location=[self.lat, self.lon], zoom_start=13)

        # Définir le rayon de recherche en mètres (10000 mètres dans cet exemple)
        rayon = 10000

        # Requête Overpass API pour récupérer les données des établissements scolaires
        overpass_url = 'http://overpass-api.de/api/interpreter'
        overpass_query = """
                [out:json];
                (
                  node["amenity"="school"](around:{}, {}, {});
                  node["amenity"="kindergarten"](around:{}, {}, {});
                );
                out body;
                >;
                out skel qt;
                """.format(rayon, self.lat, self.lon, rayon, self.lat, self.lon)

        response = requests.get(overpass_url, params={'data': overpass_query})
        data = response.json()

        # Parcourir les données et ajouter des marqueurs sur la map pour chaque établissement scolaire trouvé
        for element in data['elements']:
            if 'tags' in element:
                if 'name' in element['tags']:
                    nom = element['tags']['name']
                else:
                    nom = 'Etablissement scolaire'
                lat = element['lat']
                lon = element['lon']
                folium.Marker(
                    [lat, lon], popup=nom, icon=folium.Icon(color='blue', icon='book')
                ).add_to(map)

        # Ajouter un marqueur pour l'emplacement spécifié
        folium.Marker([self.lat, self.lon], popup='Votre emplacement').add_to(map)

        # Sauvegarder la map dans un fichier PNG
        self.path_map_etablissements_scolaires = self.__convert_map_png(
            folium_map=map, file_name='etablissements_scolaires'
        )
        return data

    @staticmethod
    def __convert_map_pdf(
        folium_map,
        file_name,
        path_to_wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe',
    ):
        if platform.system() == 'Windows':  # Windows platform
            config = pdfkit.configuration(
                wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
            )
        elif platform.system() in ['Linux', 'Darwin']:  # Linux platform
            config = pdfkit.configuration(wkhtmltopdf='/usr/local/bin/wkhtmltopdf')
        else:
            config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf)

        mapName = file_name

        # Get HTML File of Map
        folium_map.save(mapName + '.html')
        htmlfile = mapName + '.html'

        # Convert Map from HTML to PDF, Delay to Allow Rendering
        options = {
            'javascript-delay': 500,
            'page-size': 'A4',
            'margin-top': '0.0in',
            'margin-right': '0.0in',
            'margin-bottom': '0.0in',
            'margin-left': '0.0in',
            'encoding': 'UTF-8',
            'custom-header': [('Accept-Encoding', 'gzip')],
        }
        pdffile = mapName + '.pdf'

        with open(htmlfile) as f:
            pdfkit.from_file(
                input=f, output_path=pdffile, options=options, configuration=config
            )
        # Convert Map from PDF to PNG
        doc = fitz.open(pdffile)
        page = doc.load_page(0)
        pix = page.get_pixmap()
        output = mapName + '.png'
        pix.save(output)
        pngfile = mapName + '.png'
        doc.close()

        from PIL import Image

        # Crop Out Map Image
        pilImage = Image.open(pngfile)
        croppedImage = pilImage.crop(
            (0, 0, 287, 287)
        )  # Adjust this if your map renders differently on PDF

        return croppedImage

    def __convert_map_png(self, folium_map, file_name) -> FilePath:
        import io
        from PIL import Image

        htmlfile = os.path.join(f'{self.output_path}', f'{file_name}.html')
        path_pngfile = os.path.join(f'{self.output_path}', f'{file_name}.png')

        # Get HTML File of Map
        folium_map.save(htmlfile)

        img_data = folium_map._to_png(delay=2)
        pilImage = Image.open(io.BytesIO(img_data))
        pilImage.save(path_pngfile)
        # Crop Out Map Image
        croppedImage = pilImage.crop(
            (0, 0, 287, 287)
        )  # Adjust this if your map renders differently on PDF

        return path_pngfile

    def get_map_equipements_culturels_loisirs(self) -> dict:
        if self.lon is None or self.lat is None:
            self.get_localisation_du_site()

        # Créer une map centrée sur les coordonnées spécifiées
        map = folium.Map(location=[self.lat, self.lon], zoom_start=13)

        # Définir le rayon de recherche en mètres (10000 mètres dans cet exemple)
        rayon = 10000

        # Requête Overpass API pour récupérer les données des équipements culturels et de loisirs
        overpass_url = 'http://overpass-api.de/api/interpreter'
        overpass_query = """
                [out:json];
                (
                  node["amenity"="theatre"](around:{0}, {1}, {2});
                  node["amenity"="cinema"](around:{0}, {1}, {2});
                  node["leisure"="sports_centre"](around:{0}, {1}, {2});
                  node["leisure"="stadium"](around:{0}, {1}, {2});
                  node["leisure"="park"](around:{0}, {1}, {2});
                  node["leisure"="garden"](around:{0}, {1}, {2});
                  node["tourism"="museum"](around:{0}, {1}, {2});
                );
                out body;
                >;
                out skel qt;
                """.format(rayon, self.lat, self.lon)

        response = requests.get(url=overpass_url, params={'data': overpass_query})
        data = response.json()

        # Parcourir les données et ajouter des marqueurs sur la map pour chaque équipement culturel ou de loisir trouvé
        for element in data['elements']:
            if 'tags' in element:
                if 'name' in element['tags']:
                    nom = element['tags']['name']
                else:
                    nom = 'Équipement culturel ou de loisir'
                lat = element['lat']
                lon = element['lon']
                folium.Marker(
                    location=[lat, lon],
                    popup=nom,
                    icon=folium.Icon(color='green', icon='star'),
                ).add_to(map)

        # Ajouter un marqueur pour l'emplacement spécifié
        folium.Marker(location=[self.lat, self.lon], popup='Votre emplacement').add_to(
            map
        )

        # Sauvegarder la map dans un fichier HTML
        self.path_map_equipements_culturels_loisirs = self.__convert_map_png(
            folium_map=map, file_name='equipements_culturels_loisirs'
        )
        return data

    def get_map_commerces(self) -> dict:
        if self.lon is None or self.lat is None:
            self.get_localisation_du_site()

        # Créer une map centrée sur les coordonnées spécifiées
        map = folium.Map(location=[self.lat, self.lon], zoom_start=13)

        # Définir le rayon de recherche en mètres (10000 mètres dans cet exemple)
        rayon = 1000

        # Requête Overpass API pour récupérer les données des commerces
        overpass_url = 'http://overpass-api.de/api/interpreter'
        overpass_query = """
                [out:json];
                (
                  node["shop"](around:{0}, {1}, {2});
                );
                out body;
                >;
                out skel qt;
                """.format(rayon, self.lat, self.lon)

        response = requests.get(url=overpass_url, params={'data': overpass_query})
        data = response.json()

        # Parcourir les données et ajouter des marqueurs sur la map pour chaque commerce trouvé
        for element in data['elements']:
            if 'tags' in element:
                if 'name' in element['tags']:
                    nom = element['tags']['name']
                else:
                    nom = 'Commerce'
                lat = element['lat']
                lon = element['lon']
                folium.Marker(
                    [lat, lon],
                    popup=nom,
                    icon=folium.Icon(color='orange', icon='shopping-cart'),
                ).add_to(map)

        # Ajouter un marqueur pour l'emplacement spécifié
        folium.Marker(location=[self.lat, self.lon], popup='Votre emplacement').add_to(
            map
        )

        # Sauvegarder la map dans un fichier PNG
        self.path_map_commerces = self.__convert_map_png(
            folium_map=map, file_name='commerces'
        )
        return data

    def __find_nearest_point(self):
        """
        Find the nearest point in the DataFrame based on given latitude and longitude.

        Args:
        - latitude (float): The latitude of the reference point.
        - longitude (float): The longitude of the reference point.
        - dataframe (pandas.DataFrame): DataFrame containing 'lat' and 'lon' columns.

        Returns:
        - nearest_point (tuple): A tuple representing the coordinates of the nearest point.
        """
        if self.lon is None or self.lat is None:
            self.get_localisation_du_site()

        reference_point = (self.lat, self.lon)
        points = self.StationsDuDepartement[['LAT', 'LON']].values.tolist()
        distances = distance.cdist([reference_point], points)
        nearest_index = np.argmin(distances)
        self.StationsDuDepartement = self.StationsDuDepartement.reset_index()
        nearest_point = tuple(
            self.StationsDuDepartement.loc[nearest_index, ['LAT', 'LON']]
        )
        return nearest_point, nearest_index


if __name__ == '__main__':
    if platform.system() == 'Windows':
        PATH = r'C:\Users\simon\python-scripts\pymdu\TEMP'
    else:
        PATH = r'/Users/Boris/Downloads'
    diag = Diagnostic(output_path=PATH)
    diag.run()
